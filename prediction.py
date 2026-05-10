import os
import torch
import numpy as np
import imageio.v3 as iio
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import matplotlib.patches as patches
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from monai.networks.nets import Unet
from monai.networks.layers import Norm
from DataLoader import DataLoad
from monai.inferers import sliding_window_inference


model_path     = r"C:\DL\best_metric_model.pth"
save_base_dir  = r"C:\DL\result_DL"
save_xlsx_path = r"C:\DL\snr_results_DL.xlsx"

MIN_PIXEL_COUNT = 16   # 直径5px(ImageJ方式)で16px = 5.48mm²
ROI_DIAMETER    = 5    # ImageJ OvalRoi準拠: 奇数直径5px
PIXEL_AREA_MM2  = 0.5859 ** 2  # pixel_spacing=0.5859mm

# ------------------------------------------------------------------ 
#  ROI関数
# ------------------------------------------------------------------ 
def make_circular_roi(shape, center_y, center_x, diameter):
    r = diameter / 2.0
    y_grid, x_grid = np.ogrid[:shape[0], :shape[1]]
    return (x_grid - center_x) ** 2 + (y_grid - center_y) ** 2 <= r ** 2


def find_top3_rois_in_slice(slice_mask, slice_img, min_pixel_count, diameter, y_low, y_high):
    
    candidate_ys, candidate_xs = np.where(slice_mask > 0)
    if len(candidate_ys) == 0:
        return None

    mask_center = (candidate_ys >= y_low) & (candidate_ys <= y_high)
    cys = candidate_ys[mask_center]
    cxs = candidate_xs[mask_center]
    if len(cys) == 0:
        cys, cxs = candidate_ys, candidate_xs

    offsets = [0.0, 0.5]
    candidates = set()
    for cy, cx in zip(cys, cxs):
        for dy in offsets:
            for dx in offsets:
                candidates.add((cy + dy, cx + dx))

    scored = []
    for cy, cx in candidates:
        roi = make_circular_roi(slice_img.shape, cy, cx, diameter)
        roi_pixels  = int(np.sum(roi))
        inside_mask = int(np.sum(roi & (slice_mask > 0)))
        if roi_pixels == 0 or inside_mask == 0:
            continue
        signals = slice_img[roi]
        if len(signals) < 2:
            continue
        si  = float(np.mean(signals))
        sd  = float(np.std(signals))
        snr = float(si / sd) if sd > 0 else 0.0
        fully_inside = (inside_mask == roi_pixels)
        enough_size  = (roi_pixels >= min_pixel_count)
        method   = "Normal(5mm2)" if (fully_inside and enough_size) else "Fallback"
        priority = 0 if (fully_inside and enough_size) else 1
        scored.append((priority, -si, cy, cx, method, snr, si, roi))

    if not scored:
        return None

    scored.sort(key=lambda x: (x[0], x[1]))

    selected  = []
    used_mask = np.zeros(slice_img.shape, dtype=bool)
    for _, _, cy, cx, method, snr, si, roi in scored:
        if len(selected) >= 3:
            break
        if np.any(roi & used_mask):
            continue
        selected.append((cy, cx, method, snr, si, roi))
        used_mask |= roi

    return selected if selected else None


def get_best_slice_dl(pred_np, orig_np):
   
    valid_zs = [z for z in range(pred_np.shape[0]) if np.sum(pred_np[z]) > 0]
    if not valid_zs:
        return None

    all_ys       = np.concatenate([np.where(pred_np[z] > 0)[0] for z in valid_zs])
    y_min_global = int(all_ys.min())
    y_max_global = int(all_ys.max())
    y_range      = y_max_global - y_min_global
    y_low        = y_min_global + y_range * (1/3)
    y_high       = y_min_global + y_range * (2/3)

    best_avg_si  = -1
    best_z       = -1
    best_rois    = None
    best_avg_snr = -1

    for z in valid_zs:
        rois = find_top3_rois_in_slice(pred_np[z], orig_np[z], MIN_PIXEL_COUNT, ROI_DIAMETER, y_low, y_high)
        if not rois:
            continue
        avg_si = float(np.mean([r[4] for r in rois]))
        if avg_si > best_avg_si:
            best_avg_si  = avg_si
            best_z       = z
            best_rois    = rois
            best_avg_snr = float(np.mean([r[3] for r in rois]))

    if best_rois is None:
        return None
    return best_z, best_avg_snr, best_rois


def save_roi_image_dl(image_slice, rois, save_path, pt_id, cond, z, avg_snr):
    fig, ax = plt.subplots(1, 1, figsize=(6, 6))
    ax.imshow(image_slice, cmap='gray', vmin=image_slice.min(), vmax=image_slice.max())
    colors = ['red', 'blue', 'green']
    for i, (cy, cx, method, snr, si, roi) in enumerate(rois):
        coords   = np.where(roi)
        center_y = float(np.mean(coords[0]))
        center_x = float(np.mean(coords[1]))
        circle = patches.Circle(
            (center_x, center_y), radius=ROI_DIAMETER / 2,
            linewidth=0.5, edgecolor=colors[i], facecolor='none',
            label=f"ROI{i+1} SNR={snr:.1f}"
        )
        ax.add_patch(circle)
    ax.set_title(f"{pt_id} | {cond} | Z={z} | 平均SNR={avg_snr:.2f}", fontsize=10)
    ax.legend(loc='lower right', fontsize=7)
    ax.axis('off')
    plt.tight_layout()
    plt.savefig(save_path, dpi=150, bbox_inches='tight')
    plt.close(fig)


# ------------------------------------------------------------------ 
#  Excel 
# ------------------------------------------------------------------ 
HEADER_FILL = PatternFill("solid", start_color="4472C4", end_color="4472C4")
HEADER_FONT = Font(bold=True, color="FFFFFF", name="Arial", size=11)
CELL_FONT   = Font(name="Arial", size=10)
BORDER      = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'),  bottom=Side(style='thin')
)

def write_header(ws, headers, col_widths):
    ws.append(headers)
    for cell in ws[1]:
        cell.fill      = HEADER_FILL
        cell.font      = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border    = BORDER
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

def write_row(ws, row):
    ws.append(row)
    for cell in ws[ws.max_row]:
        cell.font      = CELL_FONT
        cell.alignment = Alignment(horizontal='center')
        cell.border    = BORDER


# ------------------------------------------------------------------ #
#  メイン処理
# ------------------------------------------------------------------ #
if __name__ == '__main__':
    device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

    model = Unet(spatial_dims=3, in_channels=1, out_channels=1,
                 channels=(16, 32, 64, 128, 256), strides=(2, 2, 2, 2),
                 num_res_units=2, norm=Norm.BATCH).to(device)
    model.load_state_dict(torch.load(model_path))
    model.eval()

    train_loader, val_loader, test_loader = DataLoad()
    all_loaders = {"Train": train_loader, "Val": val_loader, "Test": test_loader}

    wb = Workbook()
    wb.remove(wb.active)
    summary_data = {}  # {pt_id: {cond: avg_snr}}

    with torch.no_grad():
        for phase_name, loader in all_loaders.items():
            for i, batch_data in enumerate(loader):
                batch_size = batch_data["image"].shape[0]
                for b in range(batch_size):
                    image    = batch_data["image"][b:b+1].to(device)
                    orig_np  = batch_data["orig_image"][b, 0].cpu().numpy()
                    pt_id_raw = batch_data['pt_id']
                    cond_raw  = batch_data['condition']
                    pt_id = pt_id_raw[0] if isinstance(pt_id_raw, (list, tuple)) else pt_id_raw
                    cond  = cond_raw[0]  if isinstance(cond_raw,  (list, tuple)) else cond_raw
                    cond  = cond.replace("row ", "").strip()

                    # AI推論
                    output  = sliding_window_inference(image, (32, 128, 128), 4, model)
                    pred_np = (torch.sigmoid(output)[0, 0] > 0.4).float().cpu().numpy()

                    coords = np.where(pred_np == 1)
                    if len(coords[0]) == 0:
                        print(f"[{pt_id}-{cond}] CBD未検出")
                        continue

                    # 3ROI方式でSI最大スライスを取得
                    best = get_best_slice_dl(pred_np, orig_np)
                    if best is None:
                        print(f"[{pt_id}-{cond}] 有効ROI未発見")
                        continue

                    z, avg_snr, rois = best

                    # 画像保存
                    img_save_dir = os.path.join(save_base_dir, pt_id)
                    os.makedirs(img_save_dir, exist_ok=True)
                    img_save_path = os.path.join(img_save_dir, f"{pt_id}_{cond}_DL.png")
                    save_roi_image_dl(orig_np[z], rois, img_save_path, pt_id, cond, z, avg_snr)

                    # 患者シート取得 or 作成
                    sheet_name = pt_id[:31]
                    if sheet_name not in wb.sheetnames:
                        ws = wb.create_sheet(title=sheet_name)
                        write_header(ws,
                            ["Phase", "Condition", "Slice_Z", "ROI", "SNR", "SI", "SD", "ROI_Pixels", "ROI_mm2", "Method"],
                            [8, 12, 10, 8, 10, 10, 10, 12, 10, 16])
                        wb[sheet_name + "_avg"] if (sheet_name + "_avg") in wb.sheetnames else None
                    else:
                        ws = wb[sheet_name]

                    # 詳細表: 3ROI分
                    snr_list = []
                    for j, (cy, cx, method, snr, si, roi) in enumerate(rois, 1):
                        sd         = float(np.std(orig_np[z][roi]))
                        roi_pixels = int(np.sum(roi))
                        roi_mm2    = round(roi_pixels * PIXEL_AREA_MM2, 2)
                        write_row(ws, [phase_name, cond, z, f"ROI{j}",
                                       round(snr, 2), round(si, 2), round(sd, 2),
                                       roi_pixels, roi_mm2, method])
                        snr_list.append(snr)

                    avg_snr_val = round(float(np.mean(snr_list)), 2)

                    if pt_id not in summary_data:
                        summary_data[pt_id] = {}
                    summary_data[pt_id][cond] = avg_snr_val

                    print(f"[{pt_id}-{cond}] Z={z}  平均SNR={avg_snr_val:.2f}")

    # --- 各患者シートに平均表を追加 ---
    for pt_id, cond_dict in summary_data.items():
        sheet_name = pt_id[:31]
        if sheet_name not in wb.sheetnames:
            continue
        ws = wb[sheet_name]
        ws.append([])
        ws.append([])
        avg_row_idx = ws.max_row + 1
        for col, val in enumerate(["Condition", "SNR平均"], 1):
            cell = ws.cell(row=avg_row_idx, column=col, value=val)
            cell.fill      = HEADER_FILL
            cell.font      = HEADER_FONT
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = BORDER
        for cond, avg_snr in cond_dict.items():
            write_row(ws, [cond, avg_snr])

    # --- Summaryシート ---
    ws_sum = wb.create_sheet("Summary_DL")
    CONDITIONS = ["RG", "18on", "16on", "14on", "12on", "18off", "16off", "14off", "12off"]
    all_conds  = []
    for d in summary_data.values():
        for c in d:
            if c not in all_conds:
                all_conds.append(c)
    cond_order = [c for c in CONDITIONS if c in all_conds] + [c for c in all_conds if c not in CONDITIONS]

    write_header(ws_sum, ["PatientID"] + cond_order, [14] + [10] * len(cond_order))
    for pt_id in sorted(summary_data.keys(), key=lambda s: [int(c) if c.isdigit() else c for c in __import__('re').split(r'(\d+)', s)]):
        row = [pt_id] + [summary_data[pt_id].get(c, None) for c in cond_order]
        write_row(ws_sum, row)

    wb.save(save_xlsx_path)
    print(f"\nExcel保存完了: {save_xlsx_path}")
    print("終了")
